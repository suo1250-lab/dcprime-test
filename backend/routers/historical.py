from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import Column, Integer, String, Boolean, DateTime, ForeignKey
from sqlalchemy.sql import func
from pydantic import BaseModel
from typing import Optional, List
from database import get_db, SessionLocal
from database import Base
import models
import json
import re

router = APIRouter(prefix="/historical", tags=["historical"])


class HistoricalStudentOut(BaseModel):
    id: int
    name: str
    grade: Optional[str]
    school: Optional[str]
    subject: Optional[str]
    score: Optional[int]
    total: Optional[int]
    score_pct: Optional[int]
    outcome: Optional[str]
    source_file: Optional[str]
    question_count: int = 0

    class Config:
        from_attributes = True


class HistoricalStudentUpdate(BaseModel):
    name: str
    grade: Optional[str] = None
    school: Optional[str] = None
    subject: Optional[str] = None
    score: Optional[int] = None
    total: Optional[int] = None
    score_pct: Optional[int] = None
    outcome: Optional[str] = None


class HistoricalStudentCreate(BaseModel):
    name: str
    grade: Optional[str] = None
    school: Optional[str] = None
    subject: Optional[str] = None
    score: Optional[int] = None
    total: Optional[int] = None
    score_pct: Optional[int] = None
    outcome: str = "배정확정"
    source_file: Optional[str] = None
    question_results: Optional[dict] = None  # {"1": true/false, ...}


@router.get("", response_model=List[HistoricalStudentOut])
def list_historical(
    outcome: Optional[str] = None,
    grade: Optional[str] = None,
    subject: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.HistoricalStudent)
    if outcome:
        q = q.filter(models.HistoricalStudent.outcome == outcome)
    if grade:
        q = q.filter(models.HistoricalStudent.grade == grade)
    if subject:
        q = q.filter(models.HistoricalStudent.subject == subject)
    students = q.order_by(models.HistoricalStudent.created_at.desc()).all()
    result = []
    for s in students:
        qcount = db.query(models.HistoricalQuestionResult).filter(
            models.HistoricalQuestionResult.historical_student_id == s.id
        ).count()
        result.append(HistoricalStudentOut(
            id=s.id, name=s.name, grade=s.grade, school=s.school,
            subject=s.subject, score=s.score, total=s.total,
            score_pct=s.score_pct, outcome=s.outcome, source_file=s.source_file,
            question_count=qcount
        ))
    return result


@router.post("", status_code=201)
def create_historical(body: HistoricalStudentCreate, db: Session = Depends(get_db)):
    s = models.HistoricalStudent(
        name=body.name, grade=body.grade, school=body.school,
        subject=body.subject, score=body.score, total=body.total,
        score_pct=body.score_pct, outcome=body.outcome, source_file=body.source_file
    )
    db.add(s)
    db.flush()
    if body.question_results:
        for qno, is_correct in body.question_results.items():
            if isinstance(is_correct, bool):
                db.add(models.HistoricalQuestionResult(
                    historical_student_id=s.id,
                    question_no=int(qno),
                    is_correct=is_correct
                ))
    db.commit()
    return {"id": s.id}


@router.put("/{record_id}")
def update_historical(record_id: int, body: HistoricalStudentUpdate, db: Session = Depends(get_db)):
    s = db.query(models.HistoricalStudent).filter(models.HistoricalStudent.id == record_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    for field, val in body.model_dump(exclude_none=False).items():
        setattr(s, field, val)
    # recalculate score_pct if score/total changed
    if s.score is not None and s.total and s.total > 0:
        s.score_pct = round(s.score / s.total * 100)
    db.commit()
    return {"ok": True}


@router.delete("/{record_id}", status_code=204)
def delete_historical(record_id: int, db: Session = Depends(get_db)):
    s = db.query(models.HistoricalStudent).filter(models.HistoricalStudent.id == record_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()


@router.get("/export/excel")
def export_excel(
    outcome: Optional[str] = None,
    grade: Optional[str] = None,
    subject: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = db.query(models.HistoricalStudent)
    if outcome:
        q = q.filter(models.HistoricalStudent.outcome == outcome)
    if grade:
        q = q.filter(models.HistoricalStudent.grade == grade)
    if subject:
        q = q.filter(models.HistoricalStudent.subject == subject)
    students = q.order_by(models.HistoricalStudent.created_at.desc()).all()

    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "역대 이력"

    headers = ["이름", "학년", "학교", "과목", "점수", "만점", "점수율(%)", "결과", "파일명"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="5B6FA6")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    outcome_colors = {"배정확정": "C6EFCE", "등록불가": "FFC7CE", "포기": "D9D9D9"}
    for row_idx, s in enumerate(students, 2):
        pct = s.score_pct if s.score_pct is not None else (round(s.score / s.total * 100) if s.score and s.total else None)
        row = [s.name, s.grade, s.school, s.subject, s.score, s.total, pct, s.outcome, s.source_file]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 8 and s.outcome in outcome_colors:
                cell.fill = PatternFill("solid", fgColor=outcome_colors[s.outcome])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historical.xlsx"}
    )


# ── 인제스트 상태 (메모리) ──────────────────────────────────────
_ingest_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "skipped": 0,
    "errors": 0,
    "current": "",
    "log": [],
}

EXTRACT_PROMPT = """이 이미지는 학원 입학테스트 답안지입니다. 아래 정보를 JSON으로만 반환하세요.

추출 항목:
- name: 학생 이름
- grade: 학년 (고1/고2/고3/중1/중2/중3/초5/초6, 없으면 null)
- school: 학교명 (없으면 null)
- subject: 과목 (수학/영어/국어/과학, 없으면 null)
- score: 획득 점수 (숫자, 없으면 null)
- total: 만점 (숫자, 없으면 null)
- question_results: 문항별 정오답 {"1": true, "2": false, ...} (O=true, X=false, 표시없으면 생략)

JSON만 반환, 다른 텍스트 없이."""


def _parse_json(raw: str) -> dict:
    raw = raw.strip()
    m = re.search(r"```(?:json)?\s*([\s\S]+?)```", raw)
    if m:
        raw = m.group(1).strip()
    return json.loads(raw)


def _run_ingest():
    from pathlib import Path
    from ai_utils import ai_call
    from config import HISTORICAL_SCAN_DIR

    global _ingest_status
    _ingest_status.update({"running": True, "total": 0, "done": 0, "skipped": 0, "errors": 0, "log": []})

    folders = {
        "배정확정": "배정확정",
        "등록불가 및 포기": "등록불가",
    }

    db = SessionLocal()
    try:
        done_files = {r[0] for r in db.query(models.HistoricalStudent.source_file).filter(
            models.HistoricalStudent.source_file.isnot(None)
        ).all()}

        all_pdfs = []
        for folder_name, outcome in folders.items():
            folder = HISTORICAL_SCAN_DIR / folder_name
            if not folder.exists():
                _ingest_status["log"].append(f"⚠️ 폴더 없음: {folder}")
                continue
            for pdf in sorted(folder.glob("*.[pP][dD][fF]")):
                all_pdfs.append((pdf, outcome))

        _ingest_status["total"] = len(all_pdfs)

        for pdf_path, outcome in all_pdfs:
            fname = pdf_path.name
            _ingest_status["current"] = fname

            if fname in done_files:
                _ingest_status["skipped"] += 1
                continue

            try:
                raw = ai_call(str(pdf_path), EXTRACT_PROMPT, max_tokens=1024)
                data = _parse_json(raw)

                name = data.get("name") or fname
                grade = data.get("grade")
                school = data.get("school")
                subject = data.get("subject")
                score = data.get("score")
                total = data.get("total")
                score_pct = round(score / total * 100) if score and total else None
                q_results = data.get("question_results") or {}

                hs = models.HistoricalStudent(
                    name=name, grade=grade, school=school, subject=subject,
                    score=score, total=total, score_pct=score_pct,
                    outcome=outcome, source_file=fname
                )
                db.add(hs)
                db.flush()

                for qno, is_correct in q_results.items():
                    if isinstance(is_correct, bool):
                        db.add(models.HistoricalQuestionResult(
                            historical_student_id=hs.id,
                            question_no=int(qno),
                            is_correct=is_correct
                        ))
                db.commit()
                _ingest_status["done"] += 1
                _ingest_status["log"].append(f"✓ {name} / {grade} / {subject} / {score}/{total}")

            except Exception as e:
                db.rollback()
                _ingest_status["errors"] += 1
                _ingest_status["log"].append(f"❌ {fname}: {e}")

    finally:
        db.close()
        _ingest_status["running"] = False
        _ingest_status["current"] = ""


@router.post("/ingest", status_code=202)
def start_ingest(background_tasks: BackgroundTasks):
    if _ingest_status["running"]:
        raise HTTPException(409, "이미 실행 중입니다")
    background_tasks.add_task(_run_ingest)
    return {"message": "인제스트 시작됨"}


@router.get("/ingest/status")
def ingest_status():
    return _ingest_status


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    from sqlalchemy import func as sqlfunc
    total = db.query(models.HistoricalStudent).count()
    by_outcome = db.query(
        models.HistoricalStudent.outcome,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.outcome).all()
    by_subject = db.query(
        models.HistoricalStudent.subject,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.subject).all()
    by_grade = db.query(
        models.HistoricalStudent.grade,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.grade).all()
    return {
        "total": total,
        "by_outcome": {k: v for k, v in by_outcome if k},
        "by_subject": {k: v for k, v in by_subject if k},
        "by_grade": {k: v for k, v in by_grade if k},
    }
