import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from config import MAX_UPLOAD_EXCEL
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
from typing import List, Optional
from database import get_db
import models
from models import MathSubmission

router = APIRouter(prefix="/students", tags=["students"])


class StudentCreate(BaseModel):
    name: str
    grade: str
    school: Optional[str] = None
    class_ids: List[int] = []
    phone: Optional[str] = None
    teacher: Optional[str] = None
    historical_student_id: Optional[int] = None


class StudentOut(BaseModel):
    id: int
    name: str
    grade: str
    school: Optional[str]
    class_ids: List[int]
    class_names: List[str]
    phone: Optional[str]
    teacher: Optional[str]
    historical_student_id: Optional[int]

    class Config:
        from_attributes = True


class PatchHistorical(BaseModel):
    historical_student_id: Optional[int]


class PatchClass(BaseModel):
    class_id: Optional[int]  # null = 전체 해제, 값 = 토글(있으면 제거, 없으면 추가)


@router.get("", response_model=list[StudentOut])
def list_students(
    grade: Optional[str] = None,
    school: Optional[str] = None,
    class_id: Optional[int] = None,
    teacher: Optional[str] = None,
    name: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(models.Student).options(joinedload(models.Student.classes))
    if grade:
        q = q.filter(models.Student.grade == grade)
    if school:
        q = q.filter(models.Student.school.ilike(f"%{school}%"))
    if class_id:
        q = q.filter(models.Student.classes.any(models.Class.id == class_id))
    if teacher:
        q = q.filter(models.Student.teacher == teacher)
    if name:
        q = q.filter(models.Student.name.ilike(f"%{name}%"))
    return q.order_by(models.Student.grade, models.Student.name).all()


@router.post("", response_model=StudentOut, status_code=201)
def create_student(data: StudentCreate, db: Session = Depends(get_db)):
    student = models.Student(
        name=data.name, grade=data.grade, school=data.school,
        phone=data.phone, teacher=data.teacher,
        historical_student_id=data.historical_student_id,
    )
    if data.class_ids:
        student.classes = db.query(models.Class).filter(models.Class.id.in_(data.class_ids)).all()
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@router.put("/{student_id}", response_model=StudentOut)
def update_student(student_id: int, data: StudentCreate, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(404, "학생을 찾을 수 없습니다")
    student.name = data.name
    student.grade = data.grade
    student.school = data.school
    student.phone = data.phone
    student.teacher = data.teacher
    student.historical_student_id = data.historical_student_id
    student.classes = db.query(models.Class).filter(models.Class.id.in_(data.class_ids)).all()
    db.commit()
    db.refresh(student)
    return student


@router.patch("/{student_id}/historical")
def patch_historical(student_id: int, body: PatchHistorical, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(404, "학생을 찾을 수 없습니다")
    student.historical_student_id = body.historical_student_id
    db.commit()
    return {"ok": True}


@router.patch("/{student_id}/class")
def patch_class(student_id: int, body: PatchClass, db: Session = Depends(get_db)):
    student = db.query(models.Student).options(joinedload(models.Student.classes)).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(404, "학생을 찾을 수 없습니다")
    if body.class_id is None:
        student.classes = []
    else:
        cls = db.get(models.Class, body.class_id)
        if cls:
            if cls in student.classes:
                student.classes.remove(cls)
            else:
                student.classes.append(cls)
    db.commit()
    return {"ok": True, "class_ids": student.class_ids}


@router.delete("/{student_id}", status_code=204)
def delete_student(student_id: int, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(404, "학생을 찾을 수 없습니다")
    db.delete(student)
    db.commit()


@router.get("/{student_id}/profile")
def get_student_profile(student_id: int, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(404, "학생을 찾을 수 없습니다")

    test_results = db.query(models.TestResult).options(
        joinedload(models.TestResult.test)
    ).filter(
        models.TestResult.student_id == student_id
    ).order_by(models.TestResult.created_at.desc()).all()

    test_result_list = []
    for tr in test_results:
        test = tr.test
        test_result_list.append({
            "test_id":    tr.test_id,
            "test_title": test.title if test else "",
            "subject":    test.subject if test else "",
            "grade":      test.grade if test else "",
            "score":      tr.score,
            "total":      tr.total,
            "score_pct":  round(tr.score / tr.total * 100) if tr.total else None,
            "test_date":  str(test.test_date) if test else None,
        })

    historical_list = []
    if student.historical_student_id:
        h = db.get(models.HistoricalStudent, student.historical_student_id)
        if h:
            historical_list = [{
                "id": h.id, "subject": h.subject, "score": h.score,
                "total": h.total, "score_pct": h.score_pct,
                "outcome": h.outcome, "source_file": h.source_file,
            }]
    else:
        matches = db.query(models.HistoricalStudent).filter(
            models.HistoricalStudent.name == student.name
        ).all()
        for h in matches:
            historical_list.append({
                "id": h.id, "subject": h.subject, "score": h.score,
                "total": h.total, "score_pct": h.score_pct,
                "outcome": h.outcome, "source_file": h.source_file, "grade": h.grade,
            })

    tutoring = db.query(models.WordTutoringSession).options(
        joinedload(models.WordTutoringSession.word_test)
    ).filter(
        models.WordTutoringSession.student_id == student_id
    ).order_by(models.WordTutoringSession.session_date.desc()).limit(10).all()

    tutoring_list = []
    for t in tutoring:
        wt = t.word_test if t.word_test_id else None
        tutoring_list.append({
            "id":               t.id,
            "session_date":     str(t.session_date),
            "word_test_title":  wt.title if wt else None,
            "attempt1_total":   t.attempt1_total, "attempt1_wrong": t.attempt1_wrong,
            "attempt2_total":   t.attempt2_total, "attempt2_wrong": t.attempt2_wrong,
            "attempt3_total":   t.attempt3_total, "attempt3_wrong": t.attempt3_wrong,
            "memo":             t.memo,
        })

    # 수학 성적
    math_subs = db.query(MathSubmission).filter(
        MathSubmission.student_id == student_id,
        MathSubmission.status == "graded",
        MathSubmission.total.isnot(None),
    ).order_by(MathSubmission.submitted_at.desc()).all()

    math_list = []
    for ms in math_subs:
        # 반 평균/석차 계산
        all_subs = db.query(MathSubmission).filter(
            MathSubmission.math_test_id == ms.math_test_id,
            MathSubmission.status == "graded",
            MathSubmission.score.isnot(None),
            MathSubmission.total.isnot(None),
        ).all()
        scores = [(s.student_id, round(s.score / s.total * 100, 1)) for s in all_subs if s.total]
        avg = round(sum(p for _, p in scores) / len(scores), 1) if scores else None
        my_pct = round(ms.score / ms.total * 100, 1) if ms.total else None
        rank = sum(1 for _, p in scores if p > my_pct) + 1 if my_pct is not None and scores else None

        math_list.append({
            "id": ms.id,
            "test_title": ms.math_test.title if ms.math_test else "",
            "test_date": str(ms.math_test.test_date) if ms.math_test else None,
            "score": ms.score,
            "total": ms.total,
            "score_pct": my_pct,
            "class_avg": avg,
            "class_rank": rank,
            "class_total": len(scores),
        })

    return {
        "id":                    student.id,
        "name":                  student.name,
        "grade":                 student.grade,
        "school":                student.school,
        "phone":                 student.phone,
        "teacher":               student.teacher,
        "class_ids":             student.class_ids,
        "class_names":           student.class_names,
        "historical_student_id": student.historical_student_id,
        "test_results":          test_result_list,
        "historical":            historical_list,
        "tutoring_sessions":     tutoring_list,
        "math_results":          math_list,
    }


# ── 엑셀 일괄 import ──────────────────────────────────────────
@router.post("/import/excel")
async def import_students_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    엑셀 파일로 학생 일괄 등록/업데이트.

    엑셀 컬럼 순서 (헤더 필수):
      이름 | 학년 | 학교 | 반이름 | 전화번호 | 담당선생님

    - 반이름이 있으면 Class 테이블에서 이름+학년으로 조회, 없으면 자동 생성
    - 동일 이름+학년 학생이 있으면 업데이트, 없으면 신규 등록
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "엑셀 파일(.xlsx, .xls)만 업로드 가능합니다")
    excel_bytes = await file.read()
    if len(excel_bytes) > MAX_UPLOAD_EXCEL:
        raise HTTPException(413, "엑셀 파일이 너무 큽니다 (최대 5MB)")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl이 설치되어 있지 않습니다")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"엑셀 파일을 읽을 수 없습니다: {e}")

    # 헤더 파싱
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "파일이 비어있습니다")

    header = [str(h).strip() if h else "" for h in rows[0]]
    REQUIRED = {"이름", "학년"}
    if not REQUIRED.issubset(set(header)):
        raise HTTPException(
            400,
            f"필수 컬럼 누락: {REQUIRED - set(header)}. "
            "헤더: 이름, 학년, 학교, 반이름, 전화번호, 담당선생님"
        )

    def col(row_dict, *names):
        for n in names:
            v = row_dict.get(n)
            if v is not None:
                return str(v).strip()
        return ""

    created = 0
    updated = 0
    errors  = []

    for row_no, row_vals in enumerate(rows[1:], start=2):
        row_dict = dict(zip(header, row_vals))

        name  = col(row_dict, "이름")
        grade = col(row_dict, "학년")
        if not name or not grade:
            errors.append(f"행 {row_no}: 이름/학년 누락, 건너뜀")
            continue

        school  = col(row_dict, "학교")
        phone   = col(row_dict, "전화번호")
        teacher = col(row_dict, "담당선생님")
        class_name_raw = col(row_dict, "반이름", "반")

        # 반 조회 또는 생성
        new_cls = None
        if class_name_raw:
            new_cls = db.query(models.Class).filter(
                models.Class.name == class_name_raw,
                models.Class.grade == grade,
            ).first()
            if not new_cls:
                # 과목은 기본값 "영어" (추후 컬럼 추가 가능)
                new_cls = models.Class(name=class_name_raw, grade=grade, subject="영어")
                db.add(new_cls)
                db.flush()

        # 학생 조회 (이름+학년 기준)
        existing = db.query(models.Student).options(joinedload(models.Student.classes)).filter(
            models.Student.name  == name,
            models.Student.grade == grade,
        ).first()

        if existing:
            existing.school   = school   or existing.school
            existing.phone    = phone    or existing.phone
            existing.teacher  = teacher  or existing.teacher
            if new_cls and new_cls not in existing.classes:
                existing.classes.append(new_cls)
            updated += 1
        else:
            student = models.Student(
                name=name, grade=grade,
                school=school or None,
                phone=phone or None,
                teacher=teacher or None,
            )
            if new_cls:
                student.classes = [new_cls]
            db.add(student)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"DB 저장 실패: {e}")

    return {
        "created": created,
        "updated": updated,
        "errors":  errors,
        "total":   created + updated,
    }


@router.get("/export/excel-template")
def download_excel_template():
    """학생 일괄 등록용 엑셀 템플릿 다운로드"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "학생목록"

    headers = ["이름", "학년", "학교", "반이름", "전화번호", "담당선생님"]
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.fill       = PatternFill("solid", fgColor="2E75B6")
        cell.font       = Font(bold=True, color="FFFFFF")
        cell.alignment  = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # 예시 행
    ws.append(["홍길동", "중2", "○○중학교", "A반", "010-1234-5678", "김선생님"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"},
    )
