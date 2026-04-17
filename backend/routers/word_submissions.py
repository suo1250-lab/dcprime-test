import io
import json
import re
from pathlib import Path
from datetime import date
from logger import get_logger

log = get_logger("word_submissions")
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from limiter import limiter
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
from typing import List, Optional
from database import get_db, SessionLocal
from models import WordTest, WordSubmission, WordSubmissionItem, Student, Class
from config import GRADED, UPLOAD_DIR, KOREAN_FONT_PATH, MAX_UPLOAD_IMAGE
from ai_utils import ai_call

router = APIRouter(prefix="/word-submissions", tags=["word-submissions"])

_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*([\s\S]+?)```")

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


# ── 결과 PDF 생성 ─────────────────────────────────────────────
def _make_graded_filename_str(student_name: str, class_name: str) -> str:
    today = date.today().strftime("%Y%m%d")
    return f"{today}_{class_name}_{student_name}.pdf"


def _unique_dest(dest: Path) -> Path:
    if not dest.exists():
        return dest
    stem, suffix = dest.stem, dest.suffix
    i = 2
    while True:
        candidate = dest.parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def _get_font(doc) -> str:
    """한글 폰트 파일 경로 반환. 없으면 빈 문자열 (helv 폴백)."""
    font_path = Path(KOREAN_FONT_PATH)
    return str(font_path) if font_path.exists() else ""


def _insert_text_kr(page, pos, text: str, font: str, size: int = 11, color=(0, 0, 0)):
    """한글 폰트로 텍스트 삽입 (font = 폰트 파일 경로 문자열)"""
    if font:
        page.insert_text(pos, text, fontsize=size, fontname="korean", fontfile=font, color=color)
    else:
        page.insert_text(pos, text, fontsize=size, fontname="helv", color=color)


# === [A안] 원본 이미지만 PDF로 변환 ===
def _generate_result_pdf_a(image_path: str) -> bytes:
    """원본 이미지를 PDF 한 장으로 변환"""
    import fitz
    doc      = fitz.open()
    page     = doc.new_page(width=595, height=842)  # A4
    img_path = Path(image_path)
    if img_path.exists():
        page.insert_image(fitz.Rect(0, 0, 595, 842), filename=str(img_path))
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


# === [B안] 원본 이미지 + 채점 결과 요약 페이지 ===
def _generate_result_pdf_b(image_path: str, submission, items: list) -> bytes:
    """원본 이미지(1페이지) + 채점 결과 요약(2페이지~)"""
    import fitz

    doc  = fitz.open()
    font = _get_font(doc)

    # --- 1페이지: 원본 이미지 ---
    page1    = doc.new_page(width=595, height=842)
    img_path = Path(image_path)
    if img_path.exists():
        page1.insert_image(fitz.Rect(0, 0, 595, 842), filename=str(img_path))

    # --- 2페이지~: 채점 결과 요약 ---
    page2 = doc.new_page(width=595, height=842)
    y     = 60

    test_title = submission.word_test.title if submission.word_test else ""
    pct        = round(submission.score / submission.total * 100) if submission.total else 0
    score_clr  = (0, 0.5, 0) if pct >= 70 else (0.8, 0, 0)

    _insert_text_kr(page2, (50, y),      "채점 결과",                                  font, size=18, color=(0.1, 0.1, 0.5))
    _insert_text_kr(page2, (50, y + 35), f"학생: {submission.student_name}  |  학년: {submission.grade}", font, size=12)
    _insert_text_kr(page2, (50, y + 57), f"시험: {test_title}",                        font, size=12)
    _insert_text_kr(page2, (50, y + 79), f"점수: {submission.score} / {submission.total}  ({pct}%)", font, size=14, color=score_clr)
    y += 110

    page2.draw_line((50, y), (545, y), color=(0.7, 0.7, 0.7), width=0.8)
    y += 18

    _insert_text_kr(page2, (50, y), "문항별 결과", font, size=12, color=(0.2, 0.2, 0.2))
    y += 20

    current_page = page2
    for item in sorted(items, key=lambda x: x.item_no):
        if y > 800:
            current_page = doc.new_page(width=595, height=842)
            font = _get_font(doc)  # 새 페이지에서 폰트 재사용
            y    = 50

        is_correct  = item.is_correct
        if is_correct is True:
            mark, color = "O", (0, 0.5, 0)
        elif is_correct is False:
            mark, color = "X", (0.8, 0, 0)
        else:
            mark, color = "△", (0.8, 0.5, 0)
        student_ans = item.student_answer or "(무응답)"
        correct_ans = item.correct_answer or ""
        line        = f"{mark}  {item.item_no:2d}. {item.question}  ->  {student_ans}"
        if is_correct is not True and correct_ans:
            line += f"  (정답: {correct_ans})"

        _insert_text_kr(current_page, (60, y), line, font, size=10, color=color)
        y += 17

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


def _save_to_nas(pdf_bytes: bytes, student_name: str, class_name: str) -> str:
    """채점 결과 PDF를 NAS 채점완료 폴더에 저장. 저장 경로 반환"""
    try:
        GRADED.mkdir(parents=True, exist_ok=True)
        filename = _make_graded_filename_str(student_name, class_name)
        dest     = _unique_dest(GRADED / filename)
        dest.write_bytes(pdf_bytes)
        return str(dest)
    except Exception as e:
        log.error(f"[Submission] NAS 저장 실패: {e}")
        return ""


# ── AI 채점 (기존 유지) ───────────────────────────────────────
def grade_with_ai(image_path: str, items: list) -> list:
    try:
        # M-1: 문제/정답 값의 개행 제거하여 프롬프트 인젝션 방어
        def _sanitize(s: str) -> str:
            return str(s).replace("\n", " ").replace("\r", " ")[:100]
        answer_key = "\n".join(
            [f"{i['item_no']}. 문제: {_sanitize(i['question'])} / 정답: {_sanitize(i['answer'])}" for i in items]
        )
        prompt = f"""이 이미지는 학생이 손으로 작성한 영어 단어 시험지입니다.
아래 정답지를 참고하여 각 문항에서 학생이 쓴 답을 판독하고 채점하세요.
필체가 엉망이어도 최대한 정확하게 판독하세요.

정답지:
{answer_key}

각 문항에 대해 JSON 배열로 응답하세요. 다른 텍스트 없이 JSON만 반환하세요:
[
  {{"item_no": 1, "student_answer": "학생이 쓴 내용", "is_correct": true}},
  ...
]"""

        text = ai_call(image_path, prompt, max_tokens=2000)

        m = _JSON_FENCE_RE.search(text.strip())
        if m:
            text = m.group(1).strip()
        return json.loads(text.strip())
    except Exception as e:
        log.error(f"AI grading error: {e}")
        return []


# ── 백그라운드 AI 채점 ─────────────────────────────────────────
def _bg_grade(submission_id: int, image_path: str, items_data: list, student_name: str):
    """AI 채점 + PDF 저장을 백그라운드에서 처리"""
    db = SessionLocal()
    try:
        submission = db.query(WordSubmission).filter(WordSubmission.id == submission_id).first()
        if not submission:
            return

        ai_results = grade_with_ai(image_path, items_data)
        if not ai_results:
            return

        ai_map = {r["item_no"]: r for r in ai_results}
        score  = 0
        for item_data in items_data:
            r          = ai_map.get(item_data["item_no"], {})
            is_correct = r.get("is_correct", False)
            if is_correct:
                score += 1
            db.add(WordSubmissionItem(
                submission_id=submission_id,
                item_no=item_data["item_no"],
                question=item_data["question"],
                correct_answer=item_data["answer"],
                student_answer=r.get("student_answer", ""),
                is_correct=is_correct,
            ))
        submission.score  = score
        submission.status = "pending_review"
        db.flush()

        # NAS PDF 저장
        try:
            db_student = db.query(Student).filter(Student.name == student_name).first()
            class_name = ""
            if db_student and db_student.classes:
                class_name = db_student.classes[0].name
            saved_items = db.query(WordSubmissionItem).filter(
                WordSubmissionItem.submission_id == submission_id
            ).all()
            pdf_bytes = _generate_result_pdf_b(image_path, submission, saved_items)
            nas_path  = _save_to_nas(pdf_bytes, student_name, class_name)
            if nas_path:
                submission.image_path = nas_path
        except Exception as e:
            log.error(f"[Submission] PDF 저장 실패: {e}")

        db.commit()
        log.info(f"[Submission] 백그라운드 채점 완료: sub_id={submission_id} score={score}")
    except Exception as e:
        db.rollback()
        log.error(f"[Submission] 백그라운드 채점 실패: {e}")
    finally:
        db.close()


# ── 제출 엔드포인트 ───────────────────────────────────────────
# [NOTE] POST /word-submissions (웹 직접 제출)
# 현재 운영 방식은 NAS watcher 자동채점(watcher.py)이 메인이며,
# 이 엔드포인트는 요청에 의해 유지되지만 실제로는 거의 사용되지 않음.
#
# 채점 로직 개선(이름 매칭, 유사도 채점, 모델 변경 등) 시
# watcher.py와 이 파일(_bg_grade)이 별도로 관리되므로
# 관계없는 수정은 이 엔드포인트를 건드리지 말 것.
# 대대적인 채점 로직 변경 시에는 양쪽 모두 동기화 필요.
@router.post("", status_code=201)
@limiter.limit("120/minute")
async def submit(
    request: Request,
    background_tasks: BackgroundTasks,
    word_test_id: int   = Form(...),
    student_name: str   = Form(...),
    grade: str          = Form(...),
    image: UploadFile   = File(...),
    db: Session         = Depends(get_db),
):
    test = db.query(WordTest).filter(WordTest.id == word_test_id).first()
    if not test:
        raise HTTPException(404, "시험을 찾을 수 없습니다")

    if not UPLOAD_DIR.exists():
        raise HTTPException(503, "업로드 디렉토리에 접근할 수 없습니다 (NAS 마운트 확인 필요)")

    image_bytes = await image.read()
    if len(image_bytes) > MAX_UPLOAD_IMAGE:
        raise HTTPException(413, "이미지 파일이 너무 큽니다 (최대 20MB)")

    ext        = Path(image.filename).suffix if image.filename else ".jpg"
    submission = WordSubmission(
        word_test_id=word_test_id,
        student_name=student_name,
        grade=grade,
        status="pending_manual",
        total=len(test.items),
    )
    db.add(submission)
    db.flush()

    image_path = UPLOAD_DIR / f"sub_{submission.id}{ext}"
    image_path.write_bytes(image_bytes)
    submission.image_path = str(image_path)
    db.commit()
    db.refresh(submission)

    # AI 채점은 백그라운드에서 처리 (즉시 응답)
    items_data = [
        {"item_no": item.item_no, "question": item.question, "answer": item.answer}
        for item in test.items
    ]
    background_tasks.add_task(_bg_grade, submission.id, str(image_path), items_data, student_name)

    return {
        "id":     submission.id,
        "status": "pending_manual",
        "score":  None,
        "total":  submission.total,
        "items":  [],
    }


@router.get("/export/excel")
def export_excel(word_test_id: Optional[int] = None, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = db.query(WordSubmission).filter(WordSubmission.status == "confirmed")
    if word_test_id:
        q = q.filter(WordSubmission.word_test_id == word_test_id)
    subs = q.order_by(WordSubmission.word_test_id, WordSubmission.student_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "단어시험 결과"

    header = ["제출일", "시험명", "학년", "학생이름", "점수", "총문항", "정답률(%)"]
    if subs:
        max_items = max(len(s.items) for s in subs) if subs else 0
        for i in range(1, max_items + 1):
            header += [f"{i}번_문제", f"{i}번_학생답", f"{i}번_정답여부"]

    for col, h in enumerate(header, 1):
        cell       = ws.cell(row=1, column=col, value=h)
        cell.fill  = PatternFill("solid", fgColor="4F81BD")
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, s in enumerate(subs, 2):
        pct = round(s.score / s.total * 100) if (s.total and s.score is not None) else 0
        test_title = s.word_test.title if s.word_test else ""
        row = [s.submitted_at.strftime("%Y-%m-%d"), test_title, s.grade, s.student_name,
               s.score, s.total, pct]
        for item in sorted(s.items, key=lambda x: x.item_no):
            mark = "O" if item.is_correct is True else ("△" if item.is_correct is None else "X")
            row += [item.question, item.student_answer or "", mark]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 7:
                if pct >= 80:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif pct < 60:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=word_test_results.xlsx"},
    )


@router.get("")
def list_submissions(word_test_id: Optional[int] = None, status: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(WordSubmission).options(joinedload(WordSubmission.word_test))
    if word_test_id:
        q = q.filter(WordSubmission.word_test_id == word_test_id)
    if status:
        q = q.filter(WordSubmission.status == status)
    subs = q.order_by(WordSubmission.submitted_at.desc()).all()
    return [
        {
            "id":             s.id,
            "word_test_id":   s.word_test_id,
            "test_title":     s.word_test.title if s.word_test else "",
            "student_name":   s.student_name,
            "grade":          s.grade,
            "status":         s.status,
            "score":          s.score,
            "total":          s.total,
            "submitted_at":   s.submitted_at,
        }
        for s in subs
    ]


@router.get("/{sub_id}")
def get_submission(sub_id: int, db: Session = Depends(get_db)):
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    return {
        "id":           s.id,
        "word_test_id": s.word_test_id,
        "test_title":   s.word_test.title if s.word_test else "",
        "student_name": s.student_name,
        "grade":        s.grade,
        "status":       s.status,
        "score":        s.score,
        "total":        s.total,
        "submitted_at": s.submitted_at,
        "image_path":   s.image_path,
        "items": [
            {
                "id":             i.id,
                "item_no":        i.item_no,
                "question":       i.question,
                "correct_answer": i.correct_answer,
                "student_answer": i.student_answer,
                "is_correct":     i.is_correct,
            }
            for i in s.items
        ],
    }


class ReviewBody(BaseModel):
    items: List[dict]
    student_name: str = ""


@router.put("/{sub_id}/confirm")
def confirm_submission(sub_id: int, body: ReviewBody, db: Session = Depends(get_db)):
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    if body.student_name.strip():
        s.student_name = body.student_name.strip()
    score = 0
    for item_data in body.items:
        item = db.query(WordSubmissionItem).filter(WordSubmissionItem.id == item_data["id"]).first()
        if item:
            item.student_answer = item_data.get("student_answer", item.student_answer)
            item.is_correct     = item_data.get("is_correct", False)
            if item.is_correct:
                score += 1
    s.score  = score
    s.status = "confirmed"
    db.commit()
    db.refresh(s)

    # 확정 시 빨간펜 PDF 생성 → NAS 채점완료 폴더에 저장
    try:
        pdf_bytes = _build_marked_pdf(s)
        student   = db.query(Student).filter(Student.id == s.student_id).first()
        cls_name  = ""
        if student and student.classes:
            cls_name = student.classes[0].name
        nas_path = _save_to_nas(pdf_bytes, s.student_name, cls_name)
        if nas_path:
            s.image_path = nas_path
            db.commit()
            log.info(f"[Confirm] 빨간펜 PDF 저장: {nas_path}")
    except Exception as e:
        log.error(f"[Confirm] PDF 저장 실패 (채점은 완료됨): {e}")

    return {"status": "confirmed", "score": score, "total": s.total}


class AssignBody(BaseModel):
    student_id: int


@router.put("/{sub_id}/assign")
def assign_submission(sub_id: int, body: AssignBody, db: Session = Depends(get_db)):
    """미매칭 제출을 특정 학생에게 연결하고 confirmed 처리"""
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    student = db.query(Student).filter(Student.id == body.student_id).first()
    if not student:
        raise HTTPException(404, "Student not found")
    s.student_name = student.name
    s.grade = student.grade
    s.status = "confirmed"
    db.commit()
    return {"status": "confirmed", "student_name": student.name, "grade": student.grade}


@router.delete("/{sub_id}", status_code=204)
def delete_submission(sub_id: int, db: Session = Depends(get_db)):
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    if s.image_path:
        Path(s.image_path).unlink(missing_ok=True)
    db.delete(s)
    db.commit()


@router.put("/{sub_id}/reopen")
def reopen_submission(sub_id: int, db: Session = Depends(get_db)):
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    has_ai   = any(i.is_correct is not None for i in s.items)
    s.status = "pending_review" if has_ai else "pending_manual"
    db.commit()
    return {"status": s.status}


def _build_marked_pdf(s: WordSubmission) -> bytes:
    """빨간펜 채점 결과 PDF bytes 생성 (채점 결과만, 원본 이미지 없음)
    60문항 기준 1페이지 수용 (row_h=22)
    """
    import fitz

    doc  = fitz.open()

    # ── 채점 결과 페이지 ──────────────────────────────────────
    page  = doc.new_page(width=595, height=842)
    font  = _get_font(doc)
    test_ttl  = s.word_test.title if s.word_test else ""
    pct       = round(s.score / s.total * 100) if (s.total and s.score is not None) else 0
    wrong_nos = [i.item_no for i in s.items if i.is_correct is False]

    # 헤더
    y = 40
    _insert_text_kr(page, (50, y),      s.student_name,              font, size=15, color=(0.05, 0.05, 0.05))
    _insert_text_kr(page, (50, y + 20), f"{test_ttl}  |  {s.grade}", font, size=10, color=(0.4, 0.4, 0.4))
    score_clr = (0.8, 0, 0) if pct < 70 else (0, 0.5, 0)
    _insert_text_kr(page, (400, y), f"{s.score} / {s.total}  ({pct}%)", font, size=13, color=score_clr)

    if wrong_nos:
        wrong_str = "틀린 문항: " + ", ".join(str(n) for n in sorted(wrong_nos))
        _insert_text_kr(page, (50, y + 38), wrong_str, font, size=9, color=(0.8, 0, 0))

    y += 60
    page.draw_line((50, y), (545, y), color=(0.8, 0.1, 0.1), width=1.2)
    y += 12

    col_w   = 240
    col_gap = 15
    col_x   = [50, 50 + col_w + col_gap]
    row_h   = 22   # 60문항(30행) × 22 = 660px → 헤더 포함 832px 내 수용
    col_idx = 0

    for item in sorted(s.items, key=lambda x: x.item_no):
        cx = col_x[col_idx]
        if y > 820:
            page    = doc.new_page(width=595, height=842)
            font    = _get_font(doc)
            y       = 40
            col_idx = 0
            cx      = col_x[0]

        if item.is_correct is True:
            mark, mark_clr, text_clr = "O", (0, 0.55, 0.1), (0.15, 0.15, 0.15)
        elif item.is_correct is False:
            mark, mark_clr, text_clr = "X", (0.85, 0.05, 0.05), (0.85, 0.05, 0.05)
        else:
            mark, mark_clr, text_clr = "△", (0.75, 0.45, 0), (0.4, 0.4, 0.4)

        page.draw_circle(fitz.Point(cx + 8, y - 2), 7, color=mark_clr, fill=mark_clr if item.is_correct is False else None)
        _insert_text_kr(page, (cx + 4, y),   mark,                                      font, size=9,  color=(1,1,1) if item.is_correct is False else mark_clr)
        _insert_text_kr(page, (cx + 20, y),  f"{item.item_no}. {item.question or ''}",  font, size=9,  color=text_clr)

        if item.is_correct is False:
            stu = item.student_answer or "(무응답)"
            cor = item.correct_answer or ""
            _insert_text_kr(page, (cx + 20, y + 11), f"  {stu} → {cor}", font, size=8, color=(0.7, 0.0, 0.0))
        elif item.is_correct is True:
            _insert_text_kr(page, (cx + 20, y + 11), f"  {item.student_answer or ''}", font, size=8, color=(0.3, 0.3, 0.3))

        col_idx += 1
        if col_idx == 2:
            col_idx = 0
            y += row_h

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


@router.get("/{sub_id}/marked-pdf")
def get_marked_pdf(sub_id: int, db: Session = Depends(get_db)):
    """원본 이미지 + 빨간 펜 채점 결과 PDF 생성 및 반환"""
    from urllib.parse import quote
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s:
        raise HTTPException(404, "Not found")

    pdf_bytes = _build_marked_pdf(s)
    test_ttl  = s.word_test.title if s.word_test else ""
    fname     = f"{s.student_name}_{test_ttl}_채점결과.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"result.pdf\"; filename*=UTF-8''{quote(fname)}"},
    )


@router.get("/{sub_id}/image")
def get_image(sub_id: int, db: Session = Depends(get_db)):
    s = db.query(WordSubmission).filter(WordSubmission.id == sub_id).first()
    if not s or not s.image_path:
        raise HTTPException(404, "Image not found")
    path = Path(s.image_path)
    if not path.exists():
        raise HTTPException(404, "Image file not found")
    ext        = path.suffix.lower()
    media_map  = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png",  ".webp": "image/webp",
        ".pdf": "application/pdf",
    }
    media_type = media_map.get(ext, "image/jpeg")
    return StreamingResponse(open(path, "rb"), media_type=media_type)
