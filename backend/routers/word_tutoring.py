import json
import re
import shutil
import tempfile
from logger import get_logger

log = get_logger("word_tutoring")
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from limiter import limiter
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date
from database import get_db
import models
from ai_utils import ai_call
from config import MAX_UPLOAD_IMAGE

_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*([\s\S]+?)```")


def _grade_with_ai(image_path: str, items: list) -> list:
    try:
        answer_key = "\n".join([f"{i['item_no']}. 문제: {i['question']} / 정답: {i['answer']}" for i in items])
        prompt = f"""이 이미지는 학생이 손으로 작성한 영어 단어 시험지입니다.
아래 정답지를 참고하여 각 문항에서 학생이 쓴 답을 판독하고 채점하세요.
필체가 엉망이어도 최대한 정확하게 판독하세요.

정답지:
{answer_key}

각 문항에 대해 JSON 배열로 응답하세요. 다른 텍스트 없이 JSON만 반환하세요:
[
  {{"item_no": 1, "student_answer": "학생이 쓴 내용", "is_correct": true}},
  ...
]"""
        text = ai_call(image_path, prompt, max_tokens=2000)
        m = _JSON_FENCE_RE.search(text)
        if m:
            text = m.group(1).strip()
        return json.loads(text)
    except Exception as e:
        log.error(f"AI grading error: {e}")
        return []

router = APIRouter(prefix="/word-tutoring", tags=["word-tutoring"])


class TutoringSessionCreate(BaseModel):
    student_id: int
    word_test_id: Optional[int] = None
    session_date: date
    attempt1_total: Optional[int] = None
    attempt1_wrong: Optional[int] = None
    attempt2_total: Optional[int] = None
    attempt2_wrong: Optional[int] = None
    attempt3_total: Optional[int] = None
    attempt3_wrong: Optional[int] = None
    memo: Optional[str] = None


class TutoringSessionOut(BaseModel):
    id: int
    student_id: int
    student_name: str = ""
    word_test_id: Optional[int]
    word_test_title: Optional[str] = None
    session_date: date
    attempt1_total: Optional[int]
    attempt1_wrong: Optional[int]
    attempt2_total: Optional[int]
    attempt2_wrong: Optional[int]
    attempt3_total: Optional[int]
    attempt3_wrong: Optional[int]
    memo: Optional[str]

    class Config:
        from_attributes = True


@router.get("", response_model=List[TutoringSessionOut])
def list_sessions(student_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(models.WordTutoringSession)
    if student_id:
        q = q.filter(models.WordTutoringSession.student_id == student_id)
    sessions = q.order_by(models.WordTutoringSession.session_date.desc()).all()
    result = []
    for s in sessions:
        student = db.get(models.Student, s.student_id)
        wt = db.get(models.WordTest, s.word_test_id) if s.word_test_id else None
        result.append(TutoringSessionOut(
            id=s.id, student_id=s.student_id,
            student_name=student.name if student else "",
            word_test_id=s.word_test_id,
            word_test_title=wt.title if wt else None,
            session_date=s.session_date,
            attempt1_total=s.attempt1_total, attempt1_wrong=s.attempt1_wrong,
            attempt2_total=s.attempt2_total, attempt2_wrong=s.attempt2_wrong,
            attempt3_total=s.attempt3_total, attempt3_wrong=s.attempt3_wrong,
            memo=s.memo,
        ))
    return result


@router.get("/export/excel")
def export_excel(student_id: Optional[int] = None, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = db.query(models.WordTutoringSession)
    if student_id:
        q = q.filter(models.WordTutoringSession.student_id == student_id)
    sessions = q.order_by(models.WordTutoringSession.session_date.desc()).all()

    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "튜터링 기록"

    headers = ["날짜", "이름", "학년", "단어시험", "1차 전체", "1차 오답", "2차 전체", "2차 오답", "3차 전체", "3차 오답", "메모"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="2E7D5C")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, s in enumerate(sessions, 2):
        student = db.get(models.Student, s.student_id)
        wt = db.get(models.WordTest, s.word_test_id) if s.word_test_id else None
        row = [
            str(s.session_date),
            student.name if student else "",
            student.grade if student else "",
            wt.title if wt else "",
            s.attempt1_total, s.attempt1_wrong,
            s.attempt2_total, s.attempt2_wrong,
            s.attempt3_total, s.attempt3_wrong,
            s.memo or "",
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tutoring.xlsx"}
    )


@router.post("/grade-image")
@limiter.limit("60/minute")
async def grade_image(
    request: Request,
    word_test_id: int = Form(...),
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    test = db.query(models.WordTest).filter(models.WordTest.id == word_test_id).first()
    if not test:
        raise HTTPException(404, "시험을 찾을 수 없습니다")
    items_data = [{"item_no": i.item_no, "question": i.question, "answer": i.answer} for i in test.items]

    image_bytes = await image.read()
    if len(image_bytes) > MAX_UPLOAD_IMAGE:
        raise HTTPException(413, "이미지 파일이 너무 큽니다 (최대 20MB)")
    ext = Path(image.filename).suffix if image.filename else ".jpg"
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(image_bytes)
        tmp_path = tmp.name
    try:
        results = _grade_with_ai(tmp_path, items_data)
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if not results:
        raise HTTPException(503, "AI 채점 실패. API 키를 확인하세요.")

    total = len(results)
    correct = sum(1 for r in results if r.get("is_correct"))
    return {"total": total, "correct": correct, "wrong": total - correct, "items": results}


@router.post("", status_code=201)
def create_session(body: TutoringSessionCreate, db: Session = Depends(get_db)):
    s = models.WordTutoringSession(**body.model_dump())
    db.add(s)
    db.commit()
    return {"id": s.id}


@router.put("/{session_id}")
def update_session(session_id: int, body: TutoringSessionCreate, db: Session = Depends(get_db)):
    s = db.get(models.WordTutoringSession, session_id)
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in body.model_dump().items():
        setattr(s, k, v)
    db.commit()
    return {"ok": True}


@router.delete("/{session_id}", status_code=204)
def delete_session(session_id: int, db: Session = Depends(get_db)):
    s = db.get(models.WordTutoringSession, session_id)
    if not s:
        raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()
